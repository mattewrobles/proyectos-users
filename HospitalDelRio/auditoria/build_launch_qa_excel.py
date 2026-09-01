#!/usr/bin/env python3
"""QA de lanzamiento en producción — www.hospitaldelrio.com.ec — Excel completo."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
AZUL_OSCURO = "012274"
VERDE_OK = "C6EFCE"
AMARILLO = "FFEB9C"
ROJO = "FFC7CE"
ROJO_FUERTE = "FF5A5A"
HEADER_FILL = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fill_rows(ws, start_row, end_row, ncols, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = fill

def write_table(ws, headers, rows, start_row=2, widths=None, row_color_fn=None):
    style_header(ws, len(headers), row=start_row - 1)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row - 1, column=ci, value=h)
    r = start_row
    for row in rows:
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.alignment = WRAP
            cell.border = THIN
        if row_color_fn:
            color = row_color_fn(row)
            if color:
                fill_rows(ws, r, r, len(headers), color)
        r += 1
    if widths:
        autosize(ws, widths)
    ws.freeze_panes = f"A{start_row}"
    return r

wb = Workbook()

# ================= RESUMEN =================
ws = wb.active
ws.title = "Resumen"
ws["A1"] = "QA de Lanzamiento — www.hospitaldelrio.com.ec"
ws["A1"].font = Font(bold=True, size=16, color=AZUL_OSCURO)
ws["A2"] = "Prueba real de usuario: navegación, formularios, links, Core Web Vitals, SEO/GEO, accesibilidad. 28 de agosto de 2026."
ws["A2"].font = Font(italic=True, color="666666")

score_rows = [
    ("Core Web Vitals (datos reales de usuarios, CrUX)", ""),
    ("LCP (Largest Contentful Paint)", "1.87s — Bueno (umbral: <2.5s)"),
    ("INP (Interaction to Next Paint)", "70ms — Bueno (umbral: <200ms)"),
    ("CLS (Cumulative Layout Shift)", "0.03 — Bueno (umbral: <0.1)"),
    ("", ""),
    ("Lighthouse (Home, desktop)", ""),
    ("SEO", "100 / 100"),
    ("Best Practices", "96 / 100"),
    ("Accesibilidad", "84 / 100"),
    ("Agentic Browsing (legibilidad para IA/agentes)", "50 / 100"),
    ("", ""),
    ("HALLAZGO MÁS GRAVE", ""),
    ("Formulario de contacto NO se puede enviar", "El botón queda deshabilitado — el captcha (Cloudflare Turnstile) tira Error 600010, típico de dominio no autorizado. Nadie puede mandar una consulta por la web ahora mismo."),
    ("", ""),
    ("Otros hallazgos clave", ""),
    ("227/227 perfiles de doctores", "cargan bien (los 4 que daban 404 ya se arreglaron)"),
    ("Footer de TODO el sitio", "dice '© 2026 Hospital Universitario del Río' — mismo error que se corrigió en las descripciones, pero en el template global"),
    ("Schema.org (datos que lee Google/IA)", "el campo 'name' del Hospital dice 'Hospital Universitario del Río' — esto es lo que aparece en buscadores y respuestas de IA"),
    ("Links al sitio viejo (.webflow.io)", "siguen en el dropdown Instalaciones del nav (todas las páginas) y en la sección Servicios + Instalaciones del home"),
]
r = 4
for label, val in score_rows:
    bold = label.isupper() or label in ("Core Web Vitals (datos reales de usuarios, CrUX)", "Lighthouse (Home, desktop)", "Otros hallazgos clave")
    ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
    ws.cell(row=r, column=2, value=val)
    if "HALLAZGO" in label or "no se puede enviar" in label.lower():
        fill_rows(ws, r, r, 2, ROJO_FUERTE)
    r += 1
autosize(ws, [45, 85])

# ================= CRÍTICO =================
ws_c = wb.create_sheet("🔴 Crítico")
critico = [
    ("Formulario /contacto", "Botón 'Send message' permanece deshabilitado aunque se llenen todos los campos requeridos y se marque el checkbox de consentimiento.", "Cloudflare Turnstile (captcha) tira 'Error 600010' en consola — típicamente significa que el dominio www.hospitaldelrio.com.ec no está autorizado en la configuración del widget Turnstile (se quedó configurado solo para el dominio de staging). Revisar en el dashboard de Cloudflare Turnstile o en Webflow Site Settings > Forms.", "Se probó llenando el formulario completo con datos de prueba etiquetados 'PRUEBA QA - Cleo (ignorar)' — no se logró enviar. No se generó ningún lead real.", "URGENTE — 0 consultas pueden llegar por la web"),
    ("Footer — todas las páginas", "El copyright dice '© 2026 Hospital Universitario del Río' en every página del sitio.", "Mismo texto prohibido que se corrigió en las 177 descripciones de doctores, pero quedó en el componente global del footer — no se tocó en esa tanda porque no es parte del CMS de doctores.", "Visible en el footer de home, contacto, doctores, todas las páginas internas.", "ALTA — inconsistencia de marca en cada página"),
    ("Schema.org (JSON-LD) — Home", "El structured data del Hospital tiene \"name\": \"Hospital Universitario del Río\".", "Este dato es lo que Google usa para el Knowledge Panel y lo que leen los motores de IA (ChatGPT, Perplexity, Google AI Overviews) al responder preguntas sobre el hospital — GEO (Generative Engine Optimization). Si esto no se corrige, 'Hospital Universitario del Río' va a seguir apareciendo en resultados de búsqueda y respuestas de IA por mucho tiempo, incluso después de arreglar el resto del sitio.", "Revisado con evaluate_script en el <script type='application/ld+json'> del home.", "ALTA — afecta cómo aparece el hospital en Google/IA a futuro"),
    ("Links al dominio viejo (.webflow.io)", "El dropdown 'Instalaciones' del navbar (6 links) y las tarjetas de Instalaciones + Servicios del Home siguen apuntando a hospital-del-rio.webflow.io.", "El sitio de staging sigue público (responde 200) — cualquiera que caiga en esos links ve el sitio viejo, no el real. Ya se había avisado este riesgo antes del lanzamiento.", "El navbar es un componente global — se corrige una sola vez en Webflow y se propaga a todas las páginas.", "ALTA — riesgo de confusión/SEO duplicado"),
]
write_table(ws_c, ["Dónde", "Qué pasa", "Causa raíz / detalle técnico", "Cómo se probó", "Prioridad"], critico, widths=[22, 45, 55, 40, 22],
            row_color_fn=lambda row: ROJO_FUERTE if "URGENTE" in row[4] else ROJO)

# ================= BIEN =================
ws_b = wb.create_sheet("✅ Bien")
bien = [
    ("Core Web Vitals", "LCP 1.87s, INP 70ms, CLS 0.03 — los 3 en verde con datos reales de usuarios (CrUX), no solo de laboratorio."),
    ("SEO técnico (Lighthouse)", "100/100 — meta description, títulos, H1 y estructura básica están bien."),
    ("Directorio Médico", "227/227 perfiles de doctores cargan con status 200 — los 4 que daban 404 en la auditoría anterior ya se corrigieron."),
    ("PDFs legales", "Código de Ética y Reglamento Interno ya no apuntan al WordPress viejo — están en el CDN propio de Webflow."),
    ("Teléfonos de contacto", "6 líneas directas (Hospitalización, Consultorios, Ambulancia 24h, Farmacia, Laboratorio, Imágenes) con links tel: funcionales."),
    ("Aviso operativo visible", "El banner de 'Calle Cuzco cerrada por obras' en /contacto es un buen detalle de comunicación proactiva con el paciente."),
    ("Best Practices (Lighthouse)", "96/100 — buenas prácticas generales de desarrollo web."),
]
write_table(ws_b, ["Área", "Detalle"], bien, widths=[28, 90], row_color_fn=lambda row: VERDE_OK)

# ================= MAL (detalle) =================
ws_m = wb.create_sheet("❌ Mal — detalle")
mal = [
    ("/contacto", "Email de contacto sigue siendo un placeholder genérico: correo@hospital.com", "Contenido — no es un correo real del hospital", "ALTA"),
    ("/instalaciones-3, /instalaciones", "Ambas rutas dan 404 — el link 'Instalaciones' del footer no tiene destino válido", "Link roto — nuevo desde el lanzamiento (andaba bien en staging)", "MEDIA"),
    ("6 subpáginas de Instalaciones", "<title> genérico 'Hospital del Rio' en vez de uno específico por página (ej. 'UCI | Hospital del Río')", "SEO — mismo hallazgo de la auditoría anterior, sigue sin corregir", "MEDIA"),
    ("/lopdp-condominio", "Página con contenido de protección de datos de un CONDOMINIO, no del hospital — parece artifact del template comercial reutilizado", "Contenido fuera de lugar, ahora indexable en producción", "MEDIA"),
    ("Toda página", "Error de consola JS: \"TypeError: Cannot read properties of null (reading 'addEventListener')\" en línea 325-326 del home", "Bug de JavaScript — un script intenta engancharse a un elemento que no existe en el DOM", "MEDIA"),
    ("Toda página", "Error de consola recurrente del motor de gráficos 3D (Spline): \"Missing property\" + 54 elementos que nunca terminan de renderizar (warnings de WebGPU)", "Puede afectar contenido visual 3D — revisar en navegador real, no solo en headless", "MEDIA"),
    ("sitemap.xml", "Da 404 — no existe", "SEO — Google indexa más lento sin sitemap", "MEDIA"),
    ("robots.txt", "Existe (200) pero el archivo está vacío — no referencia ningún sitemap", "SEO — oportunidad perdida, no es un error grave por sí solo", "BAJA"),
    ("Accesibilidad — botones de video", "3 botones de play/pause de video de fondo sin nombre accesible (aria-label)", "Lectores de pantalla no anuncian qué hace el botón", "MEDIA"),
    ("Accesibilidad — links sociales del footer", "Links a Instagram/Facebook/LinkedIn/YouTube sin texto accesible (discernible name)", "Lectores de pantalla leen 'link' sin contexto", "MEDIA"),
    ("Accesibilidad — contraste de color", "Elemento con clase 'tone-medium' no cumple contraste mínimo AA", "Texto puede ser difícil de leer para baja visión", "MEDIA"),
    ("Accesibilidad — ARIA", "Elemento con atributo aria-label que además tiene texto visible distinto (label-content-name-mismatch) + uso de atributo ARIA prohibido en un heading", "Confunde a lectores de pantalla sobre qué texto leer", "BAJA"),
    ("Schema.org — geo", "El campo 'geo' usa GeoCoordinates pero solo tiene una URL de Maps, no latitude/longitude reales — no es válido según la spec de schema.org", "SEO/GEO — Google no puede usar esto para mapas/knowledge panel correctamente", "MEDIA"),
    ("Schema.org — sameAs incompleto", "Solo incluye TikTok; faltan Instagram, Facebook, LinkedIn, YouTube (todas existen y están linkeadas en el footer)", "GEO — ayuda a que Google/IA conecten todas las redes con la entidad 'Hospital del Río'", "BAJA"),
]
write_table(ws_m, ["Dónde", "Qué está mal", "Casuística técnica", "Prioridad"], mal, widths=[24, 60, 55, 12],
            row_color_fn=lambda row: ROJO if row[3] == "ALTA" else AMARILLO)

# ================= DIRECTORIO MÉDICO =================
ws_d = wb.create_sheet("Directorio Médico")
ws_d["A1"] = "227/227 perfiles de doctores probados con status HTTP real sobre www.hospitaldelrio.com.ec — todos responden 200."
ws_d["A1"].font = Font(bold=True)
ws_d["A2"] = "Los 4 doctores que en la auditoría anterior (staging) daban 404 (santiago-patricio-dominguez-vazquez, manuel-eduardo-verdugo-tapia, santiago-rafael-salamea-molina, wilson-lopez-aguirre) ya cargan correctamente en producción."
ws_d["A2"].alignment = WRAP
ws_d.merge_cells("A2:D2")
ws_d.row_dimensions[2].height = 45
autosize(ws_d, [100])

# ================= SEO / GEO / ACCESIBILIDAD =================
ws_s = wb.create_sheet("SEO · GEO · Accesibilidad")
seo_rows = [
    ("SEO técnico", "Lighthouse SEO score", "100 / 100", "OK"),
    ("SEO técnico", "Meta description (Home)", "\"Más de 100 médicos en +60 especialidades. Hospital líder del sur de Ecuador con UCI, emergencias 24/7 y tecnología avanzada.\"", "OK"),
    ("SEO técnico", "sitemap.xml", "404 — no existe", "MAL"),
    ("SEO técnico", "robots.txt", "200 pero vacío, sin referencia a sitemap", "MEJORAR"),
    ("SEO técnico", "Canonical tag (Home)", "No se encontró en el <head>", "MEJORAR"),
    ("GEO (motores de IA / respuestas generativas)", "Schema.org @type Hospital", "Presente — buena base (medicalSpecialty, availableService, amenityFeature todos poblados)", "OK con 1 problema"),
    ("GEO", "Nombre en schema.org", "\"Hospital Universitario del Río\" — debe decir \"Hospital del Río\"", "MAL — CRÍTICO"),
    ("GEO", "geo (coordenadas)", "Usa una URL de Google Maps en vez de latitude/longitude — inválido según schema.org", "MAL"),
    ("GEO", "sameAs (redes sociales)", "Solo TikTok — faltan Instagram, Facebook, LinkedIn, YouTube", "MEJORAR"),
    ("Accesibilidad", "Lighthouse Accessibility score", "84 / 100", "MEJORAR"),
    ("Accesibilidad", "Botones sin nombre accesible", "3 botones de video (play/pause)", "MAL"),
    ("Accesibilidad", "Links sin nombre accesible", "Links sociales del footer (Instagram, Facebook, LinkedIn, YouTube)", "MAL"),
    ("Accesibilidad", "Contraste de color", "1 elemento (clase 'tone-medium') no cumple mínimo AA", "MAL"),
    ("Accesibilidad", "ARIA", "1 heading con atributo ARIA prohibido + 1 mismatch entre label y texto visible", "MEJORAR"),
    ("Best Practices", "Lighthouse Best Practices score", "96 / 100", "OK"),
    ("Best Practices", "Errores de consola", "2 tipos de error recurrentes en todas las páginas (ver pestaña Mal — detalle)", "MAL"),
]
write_table(ws_s, ["Categoría", "Chequeo", "Resultado", "Estado"], seo_rows, widths=[26, 30, 60, 16],
            row_color_fn=lambda row: ROJO if "MAL" in row[3] else (AMARILLO if "MEJORAR" in row[3] else VERDE_OK))

# ================= CORE WEB VITALS =================
ws_v = wb.create_sheet("Core Web Vitals")
ws_v["A1"] = "Datos reales de usuarios (CrUX — Chrome User Experience Report), no simulados."
ws_v["A1"].font = Font(bold=True)
cwv_rows = [
    ("LCP", "Largest Contentful Paint — qué tan rápido aparece el contenido principal", "1873 ms", "< 2500 ms", "Bueno"),
    ("INP", "Interaction to Next Paint — qué tan rápido responde la página a un click/tap", "70 ms", "< 200 ms", "Bueno"),
    ("CLS", "Cumulative Layout Shift — cuánto \"salta\" el layout mientras carga", "0.03", "< 0.1", "Bueno"),
    ("TTFB", "Time to First Byte — parte del LCP, qué tan rápido responde el servidor", "645 ms", "(referencial)", "—"),
]
write_table(ws_v, ["Métrica", "Qué mide", "Valor real (p75)", "Umbral 'Bueno'", "Resultado"], cwv_rows, widths=[10, 55, 16, 16, 12],
            row_color_fn=lambda row: VERDE_OK if row[4] == "Bueno" else None)

wb._sheets = [wb["Resumen"], wb["🔴 Crítico"], wb["✅ Bien"], wb["❌ Mal — detalle"], wb["Directorio Médico"], wb["SEO · GEO · Accesibilidad"], wb["Core Web Vitals"]]

out = os.path.join(BASE, "qa-lanzamiento-produccion.xlsx")
wb.save(out)
print("Guardado:", out)
