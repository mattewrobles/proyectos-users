#!/usr/bin/env python3
"""Excel consolidado final — 225 doctores, para revisión de Mau antes de publicar."""
import json, glob, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
CMS = os.path.join(BASE, "cms")
AZUL_OSCURO = "012274"
VERDE_OK = "C6EFCE"
AMARILLO = "FFEB9C"
ROJO = "FFC7CE"
HEADER_FILL = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Cargar datos ----------
pilot = json.load(open(os.path.join(CMS, "pilot-results.json"), encoding="utf-8"))
batch_results = []
for f in sorted(glob.glob(os.path.join(CMS, "batches", "results-*.json"))):
    batch_results.extend(json.load(open(f, encoding="utf-8")))

# normalizar piloto al mismo shape que batch_results
pilot_norm = []
for d in pilot:
    redes_candidatas = []
    if d.get("redes_aprobadas"):
        for plat, url in d["redes_aprobadas"].items():
            redes_candidatas.append({"plataforma": plat, "url": url, "confianza": "APROBADA POR MAU", "fuente": "confirmado en revisión piloto"})
    pilot_norm.append({
        "slug": d["slug"],
        "item_id": d["item_id"],
        "nombre": d["nombre"],
        "descripcion_actual": d.get("descripcion_actual"),
        "descripcion_propuesta": d.get("descripcion_propuesta"),
        "cambio": d.get("cambio", ""),
        "prioridad": d.get("prioridad", ""),
        "redes_candidatas": redes_candidatas,
    })

all_doctors = pilot_norm + batch_results
print("Total doctores:", len(all_doctors))

# ---------- Números inconsistentes ----------
mismatch_path = os.path.join(CMS, "numeros-inconsistentes-final.csv")
import csv
mismatches = []
if os.path.exists(mismatch_path):
    with open(mismatch_path, encoding="utf-8") as f:
        mismatches = list(csv.DictReader(f))

wb = Workbook()

# ---------- Resumen ----------
ws = wb.active
ws.title = "Resumen"
ws["A1"] = "Ajustes CMS Directorio Médico — para tu revisión"
ws["A1"].font = Font(bold=True, size=16, color=AZUL_OSCURO)
ws["A2"] = "225 doctores con status Publish. NADA de esto se subió a Webflow todavía — es la propuesta completa, pendiente tu OK."
ws["A2"].font = Font(italic=True, color="666666")

n_desc_cambio = sum(1 for d in all_doctors if d.get("descripcion_propuesta"))
n_prioridad_alta = sum(1 for d in all_doctors if d.get("prioridad") and "ALTA" in d.get("prioridad", ""))
n_redes_candidatas = sum(len(d.get("redes_candidatas", [])) for d in all_doctors)
n_redes_alta = sum(1 for d in all_doctors for r in d.get("redes_candidatas", []) if "alta" in r.get("confianza", "").lower() or "APROBADA" in r.get("confianza", ""))

rows = [
    ("Total doctores revisados", len(all_doctors)),
    ("Descripciones con cambio propuesto", n_desc_cambio),
    ("Casos de contenido PROHIBIDO ya publicado (mención de otro hospital/'Universitario')", n_prioridad_alta),
    ("Candidatos de redes sociales encontrados (todas confianzas)", n_redes_candidatas),
    ("Candidatos de redes con confianza ALTA o ya aprobados por Mau", n_redes_alta),
    ("Doctores con número de teléfono/WhatsApp inconsistente entre campos", len(mismatches)),
    ("", ""),
    ("CÓMO REVISAR", ""),
    ("Pestaña 'Descripciones'", "Cada fila = 1 doctor. Columna F te dice qué cambié y por qué. Filas rojas = tenían contenido prohibido YA PUBLICADO."),
    ("Pestaña 'Redes candidatas'", "Redes nuevas encontradas, no publicadas. Verde = ya aprobaste (piloto). Amarillo/rojo = pendiente tu OK."),
    ("Pestaña 'Números a revisar'", "Doctores donde Teléfono/Celular/WhatsApp no coinciden entre sí — puede ser error de carga."),
]
r = 4
for label, val in rows:
    ws.cell(row=r, column=1, value=label).font = Font(bold=label.isupper() or label.startswith("Pestaña"))
    ws.cell(row=r, column=2, value=val)
    r += 1
autosize(ws, [65, 70])

# ---------- Descripciones ----------
ws_d = wb.create_sheet("Descripciones")
headers = ["Nombre", "Slug", "Item ID (Webflow)", "Descripción actual", "Descripción propuesta", "Qué cambié / por qué", "Prioridad"]
style_header(ws_d, len(headers))
for ci, h in enumerate(headers, 1):
    ws_d.cell(row=1, column=ci, value=h)
r = 2
for d in all_doctors:
    prioridad = d.get("prioridad") or ""
    desc_prop = d.get("descripcion_propuesta")
    row_vals = [
        d.get("nombre", ""), d.get("slug", ""), d.get("item_id", ""),
        (d.get("descripcion_actual") or "(vacío)"),
        (desc_prop or "(sin cambio propuesto)"),
        d.get("cambio", ""), prioridad,
    ]
    for ci, v in enumerate(row_vals, 1):
        cell = ws_d.cell(row=r, column=ci, value=v)
        cell.alignment = WRAP
        cell.border = THIN
    fill = ROJO if "ALTA" in prioridad else (AMARILLO if desc_prop else VERDE_OK)
    for ci in range(1, len(headers) + 1):
        ws_d.cell(row=r, column=ci).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    r += 1
autosize(ws_d, [26, 26, 22, 50, 50, 50, 14])
ws_d.freeze_panes = "A2"

# ---------- Redes candidatas ----------
ws_r = wb.create_sheet("Redes candidatas")
headers2 = ["Nombre", "Slug", "Plataforma", "URL", "Confianza", "Fuente / nota"]
style_header(ws_r, len(headers2))
for ci, h in enumerate(headers2, 1):
    ws_r.cell(row=1, column=ci, value=h)
r = 2
for d in all_doctors:
    for red in d.get("redes_candidatas", []):
        confianza = red.get("confianza", "")
        row_vals = [d.get("nombre", ""), d.get("slug", ""), red.get("plataforma", ""), red.get("url", ""), confianza, red.get("fuente", "")]
        for ci, v in enumerate(row_vals, 1):
            cell = ws_r.cell(row=r, column=ci, value=v)
            cell.alignment = WRAP
            cell.border = THIN
        if "APROBADA" in confianza:
            fill = VERDE_OK
        elif "alta" in confianza.lower():
            fill = VERDE_OK
        elif "baja" in confianza.lower():
            fill = ROJO
        else:
            fill = AMARILLO
        for ci in range(1, len(headers2) + 1):
            ws_r.cell(row=r, column=ci).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        r += 1
autosize(ws_r, [26, 26, 14, 45, 20, 55])
ws_r.freeze_panes = "A2"

# ---------- Números a revisar ----------
ws_n = wb.create_sheet("Números a revisar")
headers3 = ["Nombre", "Slug", "Teléfono", "Celular", "WhatsApp", "Whatsapp 2", "Detalle números distintos"]
style_header(ws_n, len(headers3))
for ci, h in enumerate(headers3, 1):
    ws_n.cell(row=1, column=ci, value=h)
r = 2
for m in mismatches:
    row_vals = [m.get("Nombre",""), m.get("Slug",""), m.get("Teléfono",""), m.get("Celular",""), m.get("WhatsApp",""), m.get("Whatsapp 2",""), m.get("Detalle números distintos","")]
    for ci, v in enumerate(row_vals, 1):
        cell = ws_n.cell(row=r, column=ci, value=v)
        cell.alignment = WRAP
        cell.border = THIN
    fill = AMARILLO
    for ci in range(1, len(headers3) + 1):
        ws_n.cell(row=r, column=ci).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    r += 1
autosize(ws_n, [26, 26, 18, 14, 45, 45, 55])
ws_n.freeze_panes = "A2"

wb._sheets = [wb["Resumen"], wb["Descripciones"], wb["Redes candidatas"], wb["Números a revisar"]]
out = os.path.join(BASE, "revision-final-cms-doctores.xlsx")
wb.save(out)
print("Guardado:", out)
