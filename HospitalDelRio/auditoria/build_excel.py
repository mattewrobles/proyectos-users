#!/usr/bin/env python3
"""Consolida la auditoría de Hospital del Río en un solo .xlsx."""
import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")

NARANJA = "FF5A00"
AZUL = "00429A"
AZUL_OSCURO = "012274"
VERDE_OK = "C6EFCE"
ROJO_MAL = "FFC7CE"
AMARILLO = "FFEB9C"

HEADER_FILL = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=AZUL_OSCURO, bold=True, size=16)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))


def style_header(ws, row=1, ncols=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, headers, rows, start_row=2, widths=None, fill_map=None):
    style_header(ws, 1, len(headers))
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, start=start_row):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = THIN
        if fill_map and fill_map(row):
            fill = fill_map(row)
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill
    if widths:
        autosize(ws, widths)
    ws.freeze_panes = "A2"


wb = Workbook()

# ---------- Cargar doctores (antes que Resumen, para poder contar) ----------
doctores = []
for f in sorted(glob.glob(os.path.join(DATA, "doctores", "lote-*.json"))):
    doctores.extend(json.load(open(f, encoding="utf-8")))

n_404 = sum(1 for d in doctores if not d.get("carga_ok", True))
n_cruzado = sum(1 for d in doctores if any(
    kw in p.lower() for p in d.get("problemas", []) for kw in ("cruzad", "otro médico", "otro doctor", "no coincide")
))
n_placeholder = sum(1 for d in doctores if any("placeholder" in p.lower() for p in d.get("problemas", [])))
n_sin_bio = sum(1 for d in doctores if d.get("carga_ok", True) and not (d.get("tiene_descripcion") and d.get("descripcion_ok")))
n_sin_whatsapp = sum(1 for d in doctores if d.get("carga_ok", True) and not d.get("tiene_whatsapp"))

# ---------- RESUMEN ----------
ws = wb.active
ws.title = "Resumen"
ws["A1"] = "Auditoría pre-lanzamiento — Hospital del Río"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Sitio: https://hospital-del-rio.webflow.io/  ·  Fecha: 2026-08-26  ·  Por: Cleo (auditoría exhaustiva)"
ws["A2"].font = Font(italic=True, color="666666")

resumen_rows = [
    ("Total páginas internas revisadas", "17 (nav, home, nosotros, planes, servicios, contacto, instalaciones+6 sub, cardiología, gastroenterología, fertilidad, dolor, legal, lopdp-condominio)"),
    ("Total doctores en Directorio Médico", "227"),
    ("Doctores con página caída (404)", str(n_404)),
    ("Doctores con contenido cruzado (bio/redes/whatsapp de OTRO doctor)", str(n_cruzado)),
    ("Doctores sin foto real (placeholder genérico)", str(n_placeholder)),
    ("Doctores sin descripción/bio", str(n_sin_bio)),
    ("Doctores sin botón WhatsApp/Agendar cita", str(n_sin_whatsapp)),
    ("Links absolutos hardcodeados a hospital-del-rio.webflow.io (riesgo migración)", "15+ (ver pestaña Riesgo Migración)"),
    ("Links apuntando al dominio de producción viejo (hospitaldelrio.com.ec)", "2 (PDFs legales en /nosotros)"),
    ("", ""),
    ("CÓMO LEER ESTE ARCHIVO", ""),
    ("✅ Bien", "Todo lo que está publicable tal como está"),
    ("❌ Mal", "Cada fila = 1 problema puntual, con link, descripción y detalle técnico"),
    ("Directorio Médico", "Los 227 doctores, uno por fila, con todos los campos auditados"),
    ("Links Rotos", "Solo los links/páginas que devuelven 404 o error real"),
    ("Riesgo Migración", "Links que se romperían al pasar de hospital-del-rio.webflow.io al dominio real"),
]
r = 4
for label, val in resumen_rows:
    ws.cell(row=r, column=1, value=label).font = Font(bold=label.isupper() or label in ("✅ Bien", "❌ Mal", "Directorio Médico", "Links Rotos", "Riesgo Migración"))
    ws.cell(row=r, column=2, value=val)
    r += 1
autosize(ws, [55, 70])

# ---------- BIEN ----------
ws_bien = wb.create_sheet("✅ Bien")
bien_rows = []
for d in doctores:
    if d.get("carga_ok") and not d.get("problemas"):
        bien_rows.append((d.get("nombre", ""), d.get("especialidad", ""), d.get("url", ""), "Perfil completo: foto real, descripción, WhatsApp/contacto, sin problemas detectados"))
# páginas institucionales sin problemas (de fase 2-5)
fase25 = json.load(open(os.path.join(DATA, "fase2-5-secciones-paginas.json"), encoding="utf-8"))
for h in fase25["hallazgos"]:
    if h["estado"] == "ok":
        bien_rows.append((h["pagina"], "", "", h["detalle"]))
fase1 = json.load(open(os.path.join(DATA, "fase1-navbar.json"), encoding="utf-8"))
for item in fase1["items"]:
    if item["estado"] == "ok":
        bien_rows.append((f"{item['seccion']} — {item['texto']}", "", item["href"], item.get("nota", "Link funcional, sintaxis correcta")))

write_rows(ws_bien, ["Doctor / Elemento", "Especialidad", "Link", "Nota"], bien_rows, widths=[38, 28, 45, 60])
for row in ws_bien.iter_rows(min_row=2, max_row=ws_bien.max_row):
    for cell in row:
        cell.fill = PatternFill(start_color=VERDE_OK, end_color=VERDE_OK, fill_type="solid")

# ---------- MAL ----------
ws_mal = wb.create_sheet("❌ Mal")
mal_rows = []

def add_mal(link, descripcion, casuistica, prioridad="MEDIA", pagina=""):
    mal_rows.append((pagina, link, descripcion, casuistica, prioridad))

# doctores con problemas
CASUISTICA_MAP = [
    ("bio", "no tiene bio", "Falta información — sin descripción/bio"),
    ("bio", "sin bio", "Falta información — sin descripción/bio"),
    ("bio", "no bio", "Falta información — sin descripción/bio"),
    ("placeholder", None, "Imagen — foto placeholder genérica (no es foto real del doctor)"),
    ("whatsapp", "sin", "Falta información — sin botón WhatsApp/Agendar cita"),
    ("agendar", "sin", "Falta información — sin botón WhatsApp/Agendar cita"),
    ("redes", "sin", "Falta información — sin redes sociales propias"),
    ("404", None, "Link roto — página no encontrada"),
    ("cruzad", None, "Contenido incorrecto — bio/datos de OTRO doctor"),
    ("otro médico", None, "Contenido incorrecto — bio/datos de OTRO doctor"),
    ("otro doctor", None, "Contenido incorrecto — bio/datos de OTRO doctor"),
    ("no coincide", None, "Inconsistencia de datos — teléfono/WhatsApp no coinciden"),
    ("wordpress", None, "Link roto / legacy — apunta al sitio viejo"),
    ("legacy", None, "Link roto / legacy — apunta al sitio viejo"),
]

def clasificar(problema):
    p = problema.lower()
    for kw, sub, cas in CASUISTICA_MAP:
        if kw in p:
            return cas
    return "Otro — revisar detalle"

for d in doctores:
    nombre = d.get("nombre") or d.get("slug", "")
    url = d.get("url", "")
    if not d.get("carga_ok", True):
        add_mal(url, f"{nombre} — página no carga (404 / Página no encontrada)", "Link roto — página no encontrada", "ALTA", "Directorio Médico")
        continue
    for problema in d.get("problemas", []):
        cas = clasificar(problema)
        prioridad = "ALTA" if ("cruzad" in problema.lower() or "otro médico" in problema.lower() or "otro doctor" in problema.lower() or "no coincide" in problema.lower()) else "MEDIA"
        add_mal(url, f"{nombre}: {problema}", cas, prioridad, "Directorio Médico")

# páginas institucionales con problemas
for h in fase25["hallazgos"]:
    if h["estado"] not in ("ok",):
        prioridad = h.get("prioridad", "MEDIA")
        add_mal(h["pagina"], h["detalle"], h["estado"], prioridad, "Página interna")

for item in fase1["items"]:
    if item["estado"] not in ("ok",):
        add_mal(item["href"], f"{item['seccion']} — {item['texto']}: {item.get('nota', item['estado'])}", item["estado"], "ALTA" if item.get("riesgo_migracion") == "ALTO" else "MEDIA", "Navbar/Footer")

# orden por prioridad
orden = {"ALTA": 0, "MEDIA": 1, "BAJA": 2}
mal_rows.sort(key=lambda r: orden.get(r[4], 1))

write_rows(ws_mal, ["Página / Sección", "Link", "Descripción del problema", "Casuística técnica", "Prioridad"], mal_rows, widths=[22, 45, 60, 40, 10])
for row in ws_mal.iter_rows(min_row=2, max_row=ws_mal.max_row):
    prioridad = row[4].value
    fill_color = ROJO_MAL if prioridad == "ALTA" else AMARILLO
    for cell in row:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# ---------- DIRECTORIO MÉDICO ----------
ws_dir = wb.create_sheet("Directorio Médico")
dir_rows = []
for d in doctores:
    dir_rows.append((
        d.get("nombre", ""),
        d.get("especialidad", ""),
        d.get("url", ""),
        "Sí" if d.get("carga_ok") else "NO — 404",
        "Sí" if d.get("tiene_foto") and d.get("foto_ok") else ("Placeholder" if d.get("tiene_foto") else "No"),
        "Sí" if d.get("tiene_descripcion") and d.get("descripcion_ok") else "No",
        "Sí" if d.get("tiene_whatsapp") else "No",
        ", ".join(d.get("redes_sociales", [])) or "—",
        len(d.get("problemas", [])),
        " | ".join(d.get("problemas", [])) or "—",
    ))
write_rows(ws_dir, ["Nombre", "Especialidad", "URL", "Carga OK", "Foto real", "Descripción", "WhatsApp", "Redes sociales", "# Problemas", "Detalle problemas"],
           dir_rows, widths=[28, 26, 42, 10, 12, 12, 10, 35, 12, 70])
for row in ws_dir.iter_rows(min_row=2, max_row=ws_dir.max_row):
    n_prob = row[8].value
    fill_color = VERDE_OK if n_prob == 0 else (ROJO_MAL if row[3].value != "Sí" else AMARILLO)
    for cell in row:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# ---------- LINKS ROTOS ----------
ws_rotos = wb.create_sheet("Links Rotos")
rotos_rows = []
for d in doctores:
    if not d.get("carga_ok", True):
        rotos_rows.append((d.get("url", ""), "Directorio Médico", "404 — Página no encontrada", "Doctor visible en /medicos pero su página individual no existe. Quitar la tarjeta del directorio o restaurar la página en Webflow."))
for h in fase25["hallazgos"]:
    if "404" in h["detalle"]:
        rotos_rows.append((h["pagina"], "Página interna", "404 referenciado", h["detalle"]))
write_rows(ws_rotos, ["Link", "Ubicación", "Tipo de error", "Detalle / Recomendación"], rotos_rows, widths=[45, 20, 25, 70])
for row in ws_rotos.iter_rows(min_row=2, max_row=ws_rotos.max_row):
    for cell in row:
        cell.fill = PatternFill(start_color=ROJO_MAL, end_color=ROJO_MAL, fill_type="solid")

# ---------- RIESGO MIGRACIÓN ----------
ws_mig = wb.create_sheet("Riesgo Migración")
ws_mig["A1"] = "Links que se romperían al migrar de hospital-del-rio.webflow.io al dominio real"
ws_mig["A1"].font = Font(bold=True, size=13, color=AZUL_OSCURO)
ws_mig["A2"] = fase1["resumen_conteo"]["recomendacion"]
ws_mig["A2"].font = Font(italic=True, color="666666")
ws_mig["A2"].alignment = WRAP
ws_mig.merge_cells("A2:E2")
ws_mig.row_dimensions[2].height = 45

mig_rows = []
for item in fase1["items"]:
    if item.get("riesgo_migracion") == "ALTO":
        mig_rows.append((item["seccion"], item["texto"], item["href"], item["tipo"], "Cambiar a link relativo en Webflow, o configurar redirect 301 wildcard al migrar dominio"))
mig_rows.append(("/nosotros", "Código de Ética / Reglamento Interno (PDF)", "https://www.hospitaldelrio.com.ec/portal/...", "absoluto a dominio de producción VIEJO", "URGENTE: subir PDFs a Webflow assets antes de dar de baja el WordPress viejo"))

write_rows(ws_mig, ["Sección", "Texto del link", "URL", "Tipo", "Recomendación"], mig_rows, start_row=4, widths=[22, 30, 55, 30, 55])
style_header(ws_mig, 4, 5)
for row in ws_mig.iter_rows(min_row=5, max_row=ws_mig.max_row):
    for cell in row:
        cell.fill = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")

# reorder sheets
wb._sheets = [wb["Resumen"], wb["✅ Bien"], wb["❌ Mal"], wb["Directorio Médico"], wb["Links Rotos"], wb["Riesgo Migración"]]

out_path = os.path.join(BASE, "hospital-del-rio-audit.xlsx")
wb.save(out_path)
print("Guardado:", out_path)
print("Doctores:", len(doctores))
print("Bien:", len(bien_rows))
print("Mal:", len(mal_rows))
print("Links rotos:", len(rotos_rows))
print("Riesgo migración:", len(mig_rows))
