#!/usr/bin/env python3
"""Excel piloto: descripciones propuestas + redes sociales candidatas (para revisión de Mau)."""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
AZUL_OSCURO = "012274"
VERDE_OK = "C6EFCE"
AMARILLO = "FFEB9C"
ROJO = "FFC7CE"
HEADER_FILL = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

data = json.load(open(os.path.join(BASE, "cms", "pilot-results.json"), encoding="utf-8"))

wb = Workbook()

# ---------- Resumen ----------
ws = wb.active
ws.title = "Resumen"
ws["A1"] = "Piloto — Ajustes CMS Directorio Médico (21 doctores)"
ws["A1"].font = Font(bold=True, size=15, color=AZUL_OSCURO)
ws["A2"] = "Muestra representativa de los 225 doctores publicados, antes de correr el resto. Nada de esto se aplicó a Webflow todavía."
ws["A2"].font = Font(italic=True, color="666666")
rows = [
    ("Descripciones corregidas ('Hospital Universitario' → 'Hospital del Río')", sum(1 for d in data if d["descripcion_actual"] and "Hospital Universitario" in d["descripcion_actual"])),
    ("Descripciones NUEVAS generadas (no tenían ninguna)", sum(1 for d in data if not d["descripcion_actual"] and d["descripcion_propuesta"])),
    ("Casos con contenido PROHIBIDO ya publicado (mención de otro hospital)", sum(1 for d in data if d.get("prioridad") == "ALTA — contenido prohibido ya publicado")),
    ("Sin cambios necesarios", sum(1 for d in data if d["descripcion_propuesta"] is None and not d.get("cambio", "").startswith("NO generé"))),
    ("Bloqueados por riesgo de tocayo/confusión — necesitan tu OK antes de escribir nada", sum(1 for d in data if d["descripcion_propuesta"] is None and d.get("cambio", "").startswith("NO generé"))),
    ("Redes sociales candidatas encontradas (sin publicar, para tu revisión)", sum(1 for d in data if d.get("redes_nota"))),
]
r = 4
for label, val in rows:
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=val)
    r += 1
autosize(ws, [70, 12])

# ---------- Descripciones ----------
ws_d = wb.create_sheet("Descripciones propuestas")
headers = ["Nombre", "Slug", "Item ID (Webflow)", "Descripción actual", "Descripción propuesta", "Qué cambié / por qué", "Prioridad"]
style_header(ws_d, len(headers))
for ci, h in enumerate(headers, 1):
    ws_d.cell(row=1, column=ci, value=h)
r = 2
for d in data:
    prioridad = d.get("prioridad", "")
    row_vals = [d["nombre"], d["slug"], d["item_id"], d["descripcion_actual"] or "(vacío)", d["descripcion_propuesta"] or "(sin cambio propuesto)", d["cambio"], prioridad]
    for ci, v in enumerate(row_vals, 1):
        cell = ws_d.cell(row=r, column=ci, value=v)
        cell.alignment = WRAP
        cell.border = THIN
    fill = ROJO if prioridad else (AMARILLO if d["descripcion_propuesta"] else VERDE_OK)
    for ci in range(1, len(headers) + 1):
        ws_d.cell(row=r, column=ci).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    r += 1
autosize(ws_d, [26, 26, 22, 55, 55, 55, 14])
ws_d.freeze_panes = "A2"

# ---------- Redes sociales candidatas ----------
ws_r = wb.create_sheet("Redes sociales candidatas")
headers2 = ["Nombre", "Slug", "Candidato(s) encontrados", "Confianza", "Acción sugerida"]
style_header(ws_r, len(headers2))
for ci, h in enumerate(headers2, 1):
    ws_r.cell(row=1, column=ci, value=h)
r = 2
for d in data:
    nota = d.get("redes_nota")
    if not nota:
        continue
    confianza = "ALTA" if "confianza ALTA" in nota else ("BAJA" if "BAJA" in nota or "RIESGO" in nota else "MEDIA")
    accion = "Verificar y aprobar" if confianza != "BAJA" else "NO usar sin confirmar directamente con el doctor"
    row_vals = [d["nombre"], d["slug"], nota, confianza, accion]
    for ci, v in enumerate(row_vals, 1):
        cell = ws_r.cell(row=r, column=ci, value=v)
        cell.alignment = WRAP
        cell.border = THIN
    fill = VERDE_OK if confianza == "ALTA" else (ROJO if confianza == "BAJA" else AMARILLO)
    for ci in range(1, len(headers2) + 1):
        ws_r.cell(row=r, column=ci).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    r += 1
autosize(ws_r, [26, 26, 75, 12, 40])
ws_r.freeze_panes = "A2"

wb._sheets = [wb["Resumen"], wb["Descripciones propuestas"], wb["Redes sociales candidatas"]]
out = os.path.join(BASE, "piloto-cms-doctores.xlsx")
wb.save(out)
print("Guardado:", out)
